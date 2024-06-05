import openpyxl

workbook = openpyxl.load_workbook("Libro.xlsx")
print(workbook.sheetnames)
print(workbook["Hoja1"])  # Objeto Worksheet

# La hoja activa será por defecto la primera, así no será necesario saber el nombre de la hoja:
hojaActiva = workbook.active
print(hojaActiva)

# CREAR HOJA NUEVA:
workbook.create_sheet("Hoja3")
print(workbook.sheetnames)

# CAMBIAR TÍTULO HOJA:
hoja3 = workbook["Hoja3"]
hoja3.title = "Nueva Titulo Hoja"
print(workbook.sheetnames)

# FILAS Y COLUMNAS:
print(f"Filas: {hojaActiva.max_row}; Columnas: {hojaActiva.max_column}")

# ACCEDER A CELDAS:
celda = hojaActiva["A1"]
print(f"\n{celda}")  # Objeto tipo Cell
print(celda.value)
celda.value = "Nombre Completo"
# Otra forma, (no parten de cero sino de 1):
celda2 = hojaActiva.cell(row=2, column=1)
print(celda2.value)
print(celda2.coordinate)


# RECORRER CELDAS:
for fila in range(1, hojaActiva.max_row+1):
    for columna in range(1, hojaActiva.max_column+1):
        celdaX = hojaActiva.cell(row=fila, column=columna)
        print(f"Celda {celdaX.coordinate}: {celdaX.value}")

# POR COLUMNA Y FILA:
columna = hojaActiva["A"]
print(f"\n{columna}")  # Tupla de objetos Cell
for celda in columna:
    print(celda.value)


# AÑADIR CELDAS:
# Nueva fila con valores 1, 2 y 3 en cada columna:
hojaActiva.append([1, 2, 3])
print(hojaActiva.rows)
for valor in hojaActiva.rows:
    print(valor)

for fila in range(1, hojaActiva.max_row+1):
    for columna in range(1, hojaActiva.max_column+1):
        celdaX = hojaActiva.cell(row=fila, column=columna)
        print(f"Celda {celdaX.coordinate}: {celdaX.value}")

# BORRAR:
# hojaActiva.delete_cols(indiceInicial, cantidad de filas/columnas a borrar)
print("Fila Borrada:")
# Borro fila Chanchito y la siguiente:
hojaActiva.delete_rows(4, 2)

for fila in range(1, hojaActiva.max_row+1):
    for columna in range(1, hojaActiva.max_column+1):
        celdaX = hojaActiva.cell(row=fila, column=columna)
        print(f"Celda {celdaX.coordinate}: {celdaX.value}")

# GUARDAR NUEVA HOJA:
workbook.save("nuevo_excel.xlsx")
